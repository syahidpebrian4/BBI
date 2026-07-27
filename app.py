import streamlit as st
import openpyxl
from openpyxl.styles import Alignment
from io import BytesIO
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# --- CONFIG ---
st.set_page_config(page_title="Draft Memo & Surat Generator", layout="wide")

# --- INISIALISASI SESSION STATE ---
if 'selected_category' not in st.session_state:
    st.session_state.selected_category = None
if 'memo_data' not in st.session_state: 
    st.session_state.memo_data = None
if 'excel_buffer' not in st.session_state: 
    st.session_state.excel_buffer = None
if 'file_name' not in st.session_state: 
    st.session_state.file_name = None

# --- CUSTOM CSS (Dominan Merah Elegant & Modern UI) ---
st.markdown("""
    <style>
    /* Styling Header / Title */
    .main-title {
        text-align: center;
        font-size: 32px;
        font-weight: 800;
        color: #b91c1c;
        margin-bottom: 4px;
        letter-spacing: -0.5px;
    }
    
    .sub-title {
        text-align: center;
        font-size: 15px;
        color: #64748b;
        margin-bottom: 35px;
    }

    /* Styling Card Utama */
    .custom-card {
        background: #ffffff;
        border: 1px solid #fee2e2;
        border-radius: 16px;
        padding: 28px 24px;
        box-shadow: 0 10px 25px -5px rgba(185, 28, 28, 0.08);
        transition: all 0.3s ease;
        position: relative;
        overflow: hidden;
    }
    
    /* Garis Aksen Kiri / Accent Bar */
    .custom-card::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 6px;
        height: 100%;
        background: linear-gradient(180deg, #dc2626 0%, #991b1b 100%);
    }

    /* Tag / Pill Badge */
    .card-tag {
        display: inline-block;
        background-color: #fef2f2;
        color: #991b1b;
        font-size: 11px;
        font-weight: 700;
        padding: 4px 10px;
        border-radius: 20px;
        margin-bottom: 12px;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        border: 1px solid #fecaca;
    }

    .card-h2 {
        font-size: 22px;
        font-weight: 700;
        color: #0f172a;
        margin-bottom: 8px;
    }

    .card-p {
        font-size: 14px;
        color: #475569;
        line-height: 1.5;
        margin-bottom: 20px;
    }

    /* Custom Styling Tombol Streamlit Utama */
    div.stButton > button {
        background: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        padding: 10px 20px !important;
        box-shadow: 0 4px 12px rgba(220, 38, 38, 0.25) !important;
        transition: all 0.2s ease !important;
    }

    div.stButton > button:hover {
        background: linear-gradient(135deg, #b91c1c 0%, #991b1b 100%) !important;
        transform: translateY(-2px);
        box-shadow: 0 6px 16px rgba(220, 38, 38, 0.35) !important;
    }
    </style>
""", unsafe_allow_html=True)

# --- FUNGSI LOGIKA NOMOR ---
def generate_memo_bbi(sheet, lokasi_transaksi):
    mapping_lokasi = {"Lotte Grosir Pasar Rebo": "01", "Lotte Grosir Kelapa Gading": "03", "Lotte Grosir Ciputat": "06"}
    kode_lokasi = mapping_lokasi.get(lokasi_transaksi, "00")
    all_rows = sheet.get_all_values()
    last_no = 0
    if len(all_rows) > 1:
        for row in reversed(all_rows[1:]):
            if row[0] and row[0].isdigit():
                last_no = int(row[0])
                break
    new_no_str = f"{last_no + 1:04d}"
    bulan_romawi = ["", "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]
    bulan = bulan_romawi[datetime.now().month]
    no_memo = f"{new_no_str}/ST{kode_lokasi}/CDHO/{bulan}/{datetime.now().year}"
    return new_no_str, no_memo

def generate_no_lapas(sheet):
    all_rows = sheet.get_all_values()
    last_no = 0
    if len(all_rows) > 1:
        for row in reversed(all_rows[1:]):
            if row[0] and row[0].isdigit():
                last_no = int(row[0])
                break
    new_no_str = f"{last_no + 1:03d}"
    bulan_romawi = ["", "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]
    bulan = bulan_romawi[datetime.now().month]
    no_surat = f"{new_no_str}/CD/LSI/{bulan}/{datetime.now().year}/E"
    return new_no_str, no_surat


# ==========================================
# 1. TAMPILAN DASHBOARD UTAMA (LANDING PAGE)
# ==========================================
if st.session_state.selected_category is None:
    st.markdown('<div class="main-title">Document Generator System</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">Pilih Jenis Dokumen yang Akan Dibuat</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2, gap="large")

    # CARD BBI
    with col1:
        st.markdown("""
            <div class="custom-card">
                <span class="card-tag">INTERNAL MEMO</span>
                <div class="card-h2">Draft Memo BBI</div>
            </div>
        """, unsafe_allow_html=True)
        st.write("")
        if st.button("Buat Memo BBI ➔", key="btn_bbi", use_container_width=True):
            st.session_state.selected_category = "BBI"
            st.session_state.memo_data = None
            st.rerun()

    # CARD LAPAS
    with col2:
        st.markdown("""
            <div class="custom-card">
                <span class="card-tag">SURAT PENAWARAN</span>
                <div class="card-h2">Draft Surat LAPAS</div>
            </div>
        """, unsafe_allow_html=True)
        st.write("")
        if st.button("Buat Surat LAPAS ➔", key="btn_lapas", use_container_width=True):
            st.session_state.selected_category = "LAPAS"
            st.session_state.memo_data = None
            st.rerun()


# ==========================================
# 2. TAMPILAN FORM INPUT (SETELAH DIPILIH)
# ==========================================
else:
    kategori = st.session_state.selected_category
    
    # Header & Tombol Kembali
    col_head1, col_head2 = st.columns([5, 1])
    with col_head1:
        st.title(f"Input Data Draft {kategori}")
    with col_head2:
        if st.button("⬅️ Ganti Menu"):
            st.session_state.selected_category = None
            st.session_state.memo_data = None
            st.rerun()

    st.markdown("---")

    # --- FORM INPUT ---
    with st.form("memo_form"):
        if kategori == "BBI":
            tanggal_input = st.date_input("Tanggal")
            no_po = st.text_input("No. PO")
            jml_artikel = st.number_input("Total jumlah artikel", min_value=0, step=1)
            
            harga_jual_str = st.text_input("Total Harga Jual Produk (contoh: 200.000.000)")
            biaya_delivery_str = st.text_input("Total Biaya Delivery (contoh: 100.000)")
            
            harga_jual = int(harga_jual_str.replace(".", "")) if harga_jual_str.replace(".", "").isdigit() else 0
            biaya_delivery = int(biaya_delivery_str.replace(".", "")) if biaya_delivery_str.replace(".", "").isdigit() else 0
            
            total_transfer = harga_jual + biaya_delivery
            st.write(f"**Total Transfer Terhitung:** {total_transfer:,}".replace(",", "."))
            
            lokasi_transaksi = st.selectbox("Lokasi transaksi", ["Lotte Grosir Pasar Rebo", "Lotte Grosir Kelapa Gading", "Lotte Grosir Ciputat"])
            rencana_transaksi = st.date_input("Rencana transaksi")

        else:  # LAPAS
            tanggal_input = st.date_input("Tanggal", value=datetime.now())
            instansi = st.text_input("Instansi (contoh: RUMAH TAHANAN NEGARA KELAS I PONDOK BAMBU)")
            lampiran = st.number_input("Jumlah Lampiran (halaman)", min_value=1, step=1, value=1)

        submitted = st.form_submit_button("Generate Draft Document")

    # --- PROSES GENERATE ---
    if submitted:
        try:
            # Koneksi Google Sheets
            creds_dict = st.secrets["gcp_service_account"]
            scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
            client = gspread.authorize(creds)
            spreadsheet = client.open("Draft Memo BBI")

            bulan_indo = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]

            if kategori == "BBI":
                sheet = spreadsheet.worksheet("BBI")
                new_no, no_memo = generate_memo_bbi(sheet, lokasi_transaksi)
                
                # Append ke Google Sheets
                row_data = [new_no, str(tanggal_input), no_memo, no_po, jml_artikel, harga_jual, biaya_delivery, total_transfer, lokasi_transaksi, str(rencana_transaksi)]
                sheet.append_row(row_data)
                
                # Process Excel Template
                wb = openpyxl.load_workbook("Draft_Memo_Template.xlsx")
                ws = wb.active

                str_tanggal = f"Jakarta, {tanggal_input.day} {bulan_indo[tanggal_input.month]} {tanggal_input.year}"
                ws['I6'] = str_tanggal
                ws['D6'] = no_memo
                ws['D8'] = no_po
                ws['F18'] = jml_artikel
                ws['G19'] = harga_jual
                ws['G20'] = biaya_delivery
                ws['G21'] = total_transfer
                ws['F22'] = lokasi_transaksi
                ws['F23'] = str(rencana_transaksi)

                for cell in ['G19', 'G20', 'G21']:
                    ws[cell].number_format = '#,##0'
                    ws[cell].alignment = Alignment(horizontal='left')

                file_prefix = "Draft Memo"

            else:  # LAPAS
                sheet = spreadsheet.worksheet("LAPAS")
                new_no, no_surat = generate_no_lapas(sheet)

                # Append ke Google Sheets
                row_data = [new_no, str(tanggal_input), no_surat, instansi, lampiran]
                sheet.append_row(row_data)

                # Process Excel Template LAPAS
                wb = openpyxl.load_workbook("Draft_LAPAS.xlsx")
                ws = wb.active

                str_tanggal = f"Jakarta, {tanggal_input.day} {bulan_indo[tanggal_input.month]} {tanggal_input.year}"
                
                ws['L6'] = str_tanggal
                ws['L6'].alignment = Alignment(horizontal='right')
                
                ws['B6'] = f': {no_surat}'
                ws['B6'].alignment = Alignment(horizontal='left')
                
                ws['B7'] = f': {lampiran} halaman'
                ws['B7'].alignment = Alignment(horizontal='left')
                
                ws['C10'] = instansi

                no_memo = no_surat
                file_prefix = "Draft LAPAS"

            output = BytesIO()
            wb.save(output)
            
            # Simpan ke session state
            st.session_state.memo_data = no_memo
            st.session_state.excel_buffer = output.getvalue()
            st.session_state.file_name = f"{file_prefix}_{new_no}.xlsx"
            st.rerun()

        except Exception as e:
            st.error(f"Terjadi kesalahan: {e}")

    # --- TAMPILAN HASIL ---
    if st.session_state.memo_data:
        st.success(f"Berhasil! Nomor Dokumen: {st.session_state.memo_data}")
        st.download_button(
            label="Download Excel", 
            data=st.session_state.excel_buffer, 
            file_name=st.session_state.file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
